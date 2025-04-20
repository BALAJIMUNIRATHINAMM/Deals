from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class DashboardFormatter:
    def __init__(self, file_path, title=None, publication_date=None):
        self.workbook = load_workbook(file_path)
        self.file_path = file_path
        self.title = title
        self.publication_date = publication_date

    def _set_title(self, sheet, cell_address):
        sheet[cell_address] = self.title
        sheet[cell_address].font = Font(name="Calibri", size=10, italic=True)

    def _paste_dataframe(self, sheet, df, start_row, start_col):
        font_style = Font(name="Calibri", size=10)
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font_style

    def _align_columns(self, sheet, center_cols, left_cols, start_row):
        for col in center_cols:
            for cell in sheet[col]:
                if cell.row >= start_row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in left_cols:
            for cell in sheet[col]:
                if cell.row >= start_row:
                    cell.alignment = Alignment(horizontal="left")

    def _adjust_column_width(self, sheet, columns, padding=0):
        for col in columns:
            max_length = max((len(str(cell.value)) for cell in sheet[col] if cell.value), default=0)
            sheet.column_dimensions[col].width = max_length + padding

    def _set_borders(self, sheet, start_row, start_col, num_rows, num_cols):
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        medium_border_left = Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
        medium_border_right = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))

        end_row = start_row + num_rows - 1
        for row in range(start_row, end_row + 1):
            for col in range(start_col, start_col + num_cols):
                cell = sheet.cell(row=row, column=col)
                cell.border = border

            sheet.cell(row=row, column=start_col).border = medium_border_left
            sheet.cell(row=row, column=start_col + num_cols - 1).border = medium_border_right

        for col in range(start_col, start_col + num_cols):
            bottom_cell = sheet.cell(row=end_row, column=col)
            bottom_cell.border = Border(
                left=bottom_cell.border.left,
                right=bottom_cell.border.right,
                top=bottom_cell.border.top,
                bottom=Side(style='medium')
            )

    def _center_align_dash_cells(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "-":
                    cell.alignment = Alignment(horizontal="center")

    def _process_sheet(self, sheet_name, df, title_cell, start_row, start_col, center_cols, left_cols, width_cols):
        sheet = self.workbook[sheet_name]
        self._set_title(sheet, title_cell)
        self._paste_dataframe(sheet, df, start_row, start_col)
        self._align_columns(sheet, center_cols, left_cols, start_row)
        self._adjust_column_width(sheet, width_cols, padding=2)
        self._set_borders(sheet, start_row, start_col, len(df), len(df.columns))
        self._center_align_dash_cells(sheet)

    def internal_deals(self, df):
        self._process_sheet(
            sheet_name="Outsourcing Dashboard",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3, 16)],
            left_cols=['E', 'J', 'N', 'O', 'P'],
            width_cols=['L']
        )

    def client_deals(self, df):
        self._process_sheet(
            sheet_name="deals",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3, 13)],
            left_cols=['E', 'H', 'K', 'L', 'M'],
            width_cols=['L']
        )

    def zinnov_deals(self, df):
        self._process_sheet(
            sheet_name="zinnov",
            df=df,
            title_cell="E3",
            start_row=6,
            start_col=7,
            center_cols=[get_column_letter(i) for i in range(7, 22)],
            left_cols=['I', 'N', 'S', 'T', 'U', 'V'],
            width_cols=['J']
        )
    def format_digital_initiatives(self, df):
        self._process_sheet(
            sheet_name="Digital Initiatives",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3, 12)],
            left_cols=['E', 'G','H','K'],
            width_cols=['F']
        )
    def format_techstack(self, df):
        self._process_sheet(
            sheet_name="Techstack Dashboard",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3, 8)],
            left_cols=[],
            width_cols=[]
        )
    def format_hiring(self, df):
        self._process_sheet(
            sheet_name="Hiring Details",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3, 15)],
            left_cols=['E', 'H' ,'I','J','K','L','M','N'],
            width_cols=['F']
        )
    def format_Keysignal(self, df):
        self._process_sheet(
            sheet_name="Key Signals",
            df=df,
            title_cell="C3",
            start_row=6,
            start_col=3,
            center_cols=[get_column_letter(i) for i in range(3,9)],
            left_cols=['E', 'F','G'],
            width_cols=['H','I']
        )


    def save(self, output_path):
        self.workbook.save(output_path)
